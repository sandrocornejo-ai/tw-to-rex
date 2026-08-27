#!/usr/bin/env python3
"""
tw_to_rex_app.py — Aplicación Streamlit
Transforma liquidaciones de TeamWork al formato de importación Rex+.

Uso:
    streamlit run tw_to_rex_app.py
"""

import io
import re
import openpyxl
import streamlit as st

# ─────────────────────────────────────────────────────────────────
#  CONSTANTES / REGLAS DE NEGOCIO
# ─────────────────────────────────────────────────────────────────

HABERES_NEGATIVOS = {
    'DESCTO. HORAS ATRASO',
    'DESCTO. HORAS ATRASO PT',
    'HORAS NO TRABAJADAS $',
    'DESC. PAGO EN EXC. IMPONIBLE',
}

EXCLUIR_DE_EXENTOS = set()   # SOBREGIRO (compensaSobre) sí entra en rentas no gravadas

SIEMPRE_GENERAR = {'impuesto', 'cesEmpleado'}

MESES_ES = {
    'ENERO': '01', 'FEBRERO': '02', 'MARZO': '03', 'ABRIL': '04',
    'MAYO': '05', 'JUNIO': '06', 'JULIO': '07', 'AGOSTO': '08',
    'SEPTIEMBRE': '09', 'OCTUBRE': '10', 'NOVIEMBRE': '11', 'DICIEMBRE': '12',
}

CONCEPTOS_AFP_INST = {
    'afp', 'aporteAFPemp', 'reliquidaAporteAFP', 'cesEmpleado',
    'reliquidaCesSol', 'cesAporteSol', 'reliquidaCesCi', 'cesAporteCi',
    'reliquidaCesEmpl', 'reliquidaAfp', 'afpAhor', 'reliquidaTrabPesa',
    'reliquidaTrabEmpl', 'trabajoPesaEmpl', 'trabajoPesa',
}

CONCEPTOS_CCAF = {
    'reliquidaCcaf', 'cajaComp', 'cajaSegu', 'cajaOtro', 'cajaVida',
    'cajaLeas', 'cajaDent', 'cajaCred', 'ccafReliquida', 'cajaAhor',
}

CONCEPTOS_MUTUAL = {'mutual', 'reliquidaMutual'}

CONCEPTOS_SEGURIDADSOCIAL = {'aporteFAPPBAC', 'aporteFAPPCEV', 'reliquidaAporteBAC'}

ISAPRE_MAPPING = {
    'BANMEDICA': 'banmedica',
    'COLMENA GOLDEN CROSS': 'colmena',
    'CONSALUD': 'consalud',
    'CRUZ BLANCA': 'cruzblanca',
    'ESENCIAL': 'esencial',
    'ISALUD': 'isalud',
    'NUEVA MASVIDA': 'nuevamasvida',
    'VIDA TRES': 'vidatres',
}

CONCEPTOS_BASE_AFP = {
    'afp', 'afpAhor', 'reliquidaAfp', 'aporteAFPemp', 'reliquidaAporteAFP',
    'aporteFAPPBAC', 'aporteFAPPCEV', 'mutual', 'reliquidaMutual',
    'sis', 'trabajoPesa', 'reliquidaTrabPesa', 'trabajoPesaEmpl', 'reliquidaTrabEmpl',
    'isapre', 'reliquidaIsapre',
}
CONCEPTOS_BASE_CES = {
    'cesEmpleado', 'cesAporteSol', 'cesAporteCi',
    'reliquidaCesSol', 'reliquidaCesCi', 'reliquidaCesEmpl', 'solidarioremu',
}
CONCEPTOS_BASE_IMP = {
    'impuesto', 'impuestoAgricola', 'reliquidaImpuesto', 'imptoindMi',
}
CONCEPTOS_BASE_TOT = {'totalesEmpl'}

DESC_LEGAL_MANUALES = {
    'AFP', 'FONASA', 'ISAPRE', 'IMPUESTO UNICO',
    'IMPUESTO UNICO DOBLE CONTRATO', 'SEG SES TRAB',
    'APV', 'APV REG (A)', 'OTROS APV', 'LIQUIDO', 'SUELDO DEL MES',
}

OUTPUT_HEADERS = [
    'Fecha de proceso', 'Id empleado', 'Número de contrato', 'Id del concepto',
    'Monto del concepto', 'Afecto', 'Id de institución', 'Cotización de jubilación',
    'Días de licencias', 'Días trabajados', 'Fecha de aplicación', 'Empresa',
    'Total de rebajas por LLSS', 'Rentas no gravadas', 'Rebaja por zona extrema',
    'Jornada', 'Días de vacaciones', 'Monto Init', 'Fase', 'Parcial7', 'Parcial8',
]


# ─────────────────────────────────────────────────────────────────
#  FUNCIONES AUXILIARES
# ─────────────────────────────────────────────────────────────────

def n(v):
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(',', '.').strip())
    except Exception:
        return 0.0


def formatear_rut(raw):
    if not raw:
        return ''
    s = str(raw).strip().replace('.', '')
    partes = s.split('-')
    if len(partes) == 2:
        return f"{partes[0].lstrip('0')}-{partes[1]}"
    return s


def parsear_periodo(texto):
    upper = texto.upper() if texto else ''
    for nombre, num in MESES_ES.items():
        if nombre in upper:
            m = re.search(r'(\d{4})', texto)
            if m:
                return f"{m.group(1)}-{num}"
    return None


def desde_agosto_2026(periodo):
    try:
        y, mo = map(int, periodo.split('-'))
        return (y > 2026) or (y == 2026 and mo >= 8)
    except Exception:
        return False


def get_afecto(id_concepto, base_afp, base_ces, base_imp, suma_afectos):
    if id_concepto in CONCEPTOS_BASE_AFP:
        return round(base_afp)
    if id_concepto in CONCEPTOS_BASE_CES:
        return round(base_ces)
    if id_concepto in CONCEPTOS_BASE_IMP:
        return round(base_imp)
    if id_concepto in CONCEPTOS_BASE_TOT:
        return round(suma_afectos)
    return 0


# ─────────────────────────────────────────────────────────────────
#  FUNCIONES DE CARGA DE ARCHIVOS (desde bytes en memoria)
# ─────────────────────────────────────────────────────────────────

def cargar_equivalencias(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active
    equiv = {}
    for row in ws.iter_rows(values_only=True):
        if not row[0] or row[0] == 'Nombre columna':
            continue
        col_name    = str(row[0]).strip()
        id_concepto = str(row[1]).strip() if row[1] else ''
        tipo        = str(row[2]).strip() if row[2] else ''
        obs         = str(row[3]).strip() if row[3] else ''
        equiv[col_name] = {'id_concepto': id_concepto, 'tipo': tipo, 'obs': obs}
    wb.close()
    return equiv


def cargar_empleados(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    headers = [str(h).strip() if h else '' for h in rows[1]]
    empleados = {}
    for row in rows[2:]:
        if not row[0]:
            continue
        d = {headers[i]: row[i] for i in range(len(headers))}
        key = str(d.get('Nombre del contrato', '')).strip()
        if key:
            empleados[key] = d
    return empleados


def cargar_params(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdrs = [str(h).strip() if h else '' for h in rows[0]]
    params = {}
    for row in rows[1:]:
        if not row[0]:
            continue
        key = str(row[0]).strip()
        params[key] = {hdrs[i]: row[i] for i in range(len(hdrs))}
    return params


def cargar_cot_afp(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active
    cot_afp = {}
    for row in ws.iter_rows(values_only=True):
        if not row[0] or row[0] == 'id_afp_hist':
            continue
        cot_afp[str(row[0]).strip()] = n(row[4])
    wb.close()
    return cot_afp


def cargar_asig(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    asig = {}
    for row in rows[1:]:
        if not row[0]:
            continue
        asig[str(row[0]).strip()] = {
            'hasta_jul': str(row[1]).strip() if row[1] else '',
            'desde_ago': str(row[2]).strip() if row[2] else '',
        }
    return asig


def cargar_tw(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


# ─────────────────────────────────────────────────────────────────
#  LÓGICA PRINCIPAL DE TRANSFORMACIÓN
# ─────────────────────────────────────────────────────────────────

def procesar(tw_bytes, equiv_bytes, emp_bytes, params_bytes, cot_bytes, asig_bytes=None,
             progress_callback=None, log_callback=None):
    """
    Ejecuta la transformación completa y retorna (output_rows, advertencias, stats).
    progress_callback(fraccion): actualiza barra de progreso (0.0 a 1.0)
    log_callback(texto): emite mensajes de estado
    """

    def log(msg):
        if log_callback:
            log_callback(msg)

    # ── Cargar referencias ──────────────────────────────────────────
    log("Cargando Equivalencias Tw.xlsx…")
    equiv = cargar_equivalencias(equiv_bytes)
    log(f"  ✓ {len(equiv)} equivalencias")

    log("Cargando empleadostw.xlsx…")
    empleados = cargar_empleados(emp_bytes)
    log(f"  ✓ {len(empleados)} empleados")

    log("Cargando parametrosMesuales.xlsx…")
    params = cargar_params(params_bytes)
    log(f"  ✓ {len(params)} períodos")

    log("Cargando cot_afp_hist.xlsx…")
    cot_afp = cargar_cot_afp(cot_bytes)
    log(f"  ✓ {len(cot_afp)} cotizaciones AFP")

    if asig_bytes:
        log("Cargando Asig Inst LD.xlsx…")
        asig = cargar_asig(asig_bytes)
        log(f"  ✓ {len(asig)} reglas de institución")
    else:
        asig = {}
        log("  ℹ Asig Inst LD.xlsx no cargado (opcional)")

    # ── Funciones que dependen de las tablas cargadas ───────────────
    def get_institucion(concepto, emp, row, desde_ago):
        if concepto in CONCEPTOS_AFP_INST:
            nombre_afp = str(safe_val(row, IDX_NOMBRE_AFP) or '').strip().lower()
            return nombre_afp if nombre_afp else 'afp'
        if concepto in CONCEPTOS_CCAF:
            return emp.get('Id CCAF') or ''
        if concepto in CONCEPTOS_MUTUAL:
            return emp.get('Id Mutual') or ''
        if concepto == 'isapre':
            inst = str(safe_val(row, IDX_INST_SALUD) or '').strip().upper()
            if not inst:
                return 'fonasa'
            return ISAPRE_MAPPING.get(inst, 'Falta Id Salud')
        if concepto == 'impuesto':
            return 'Impuesto'
        if concepto == 'sis':
            if desde_ago:
                return 'seguridadsocial'
            return str(safe_val(row, IDX_NOMBRE_AFP) or '').strip().lower()
        if concepto in CONCEPTOS_SEGURIDADSOCIAL:
            return 'seguridadsocial'
        # fallback: usar tabla Asig Inst LD
        regla = asig.get(concepto, {})
        col = regla.get('desde_ago' if desde_ago else 'hasta_jul', '')
        if col == 'Id afp':
            return emp.get('Id Afp') or ''
        if col == 'Id Salud':
            return emp.get('Id Salud') or ''
        if col == 'Id Mutual':
            return emp.get('Id Mutual') or ''
        if col == 'Id CCAF':
            return emp.get('Id CCAF') or ''
        if col == 'seguridadsocial':
            return 'seguridadsocial'
        if col == 'Impuesto':
            return 'Impuesto'
        return None

    def get_cotizacion(concepto, emp, afecto, periodo, desde_ago,
                       sis_tasa, aporte_afp, seg_vida, aporte_bac):
        id_afp = emp.get('Id Afp', '') or ''
        if concepto == 'afp':
            key = f"{periodo}{id_afp}"
            cot = cot_afp.get(key, 0)
            if not cot and '-' in periodo:
                y, mo = int(periodo[:4]), int(periodo[5:])
                mo -= 1
                if mo == 0:
                    mo, y = 12, y - 1
                key_prev = f"{y:04d}-{mo:02d}{id_afp}"
                cot = cot_afp.get(key_prev, 0)
            return round(cot * 100, 4) if cot else None
        if concepto == 'isapre':
            return afecto
        if concepto == 'mutual':
            return n(emp.get('% Mutual', 0))
        if concepto == 'sis':
            return sis_tasa
        if concepto == 'aporteAFPemp':
            return aporte_afp
        if concepto == 'aporteFAPPCEV':
            return seg_vida
        if concepto == 'aporteFAPPBAC':
            return aporte_bac
        if concepto == 'totalesEmpl':
            return afecto
        if concepto == 'cesEmpleado':
            return 0.6
        return None

    # ── Cargar tw.xlsx ──────────────────────────────────────────────
    log("Cargando tw.xlsx…")
    tw_rows_all = cargar_tw(tw_bytes)

    periodo = parsear_periodo(str(tw_rows_all[1][0]))
    if not periodo:
        raise ValueError("No se pudo detectar el período en la fila 2 de tw.xlsx")
    log(f"  ✓ Período detectado: {periodo}")

    tw_hdrs = [str(h).strip() if h is not None else '' for h in tw_rows_all[7]]
    tw_data  = tw_rows_all[8:]
    log(f"  ✓ {len(tw_data)} empleados en tw.xlsx")

    # ── Parámetros del período ──────────────────────────────────────
    pm = params.get(periodo, {})
    if not pm:
        log(f"  ⚠ Sin parámetros para {periodo}, usando ceros")

    tope_afp   = n(pm.get('topeImp_pesos_afp', 0))
    tope_ces   = n(pm.get('topeCes_pesos', 0))
    tope_salud = n(pm.get('topeSalud_pesos', 0))
    sis_tasa   = n(pm.get('sis', 0))
    aporte_afp = n(pm.get('Aporte AFP', 0))
    seg_vida   = n(pm.get('Seg Social Exp vida', 0))
    aporte_bac = n(pm.get('aporteFAPPBAC', 0))
    aporte_ccaf = n(pm.get('aporte_Ccaf', 0))
    desde_ago  = desde_agosto_2026(periodo)

    log(f"  Tope AFP: {tope_afp:,.0f}  |  Tope CES: {tope_ces:,.0f}  |  Tope Salud: {tope_salud:,.0f}")

    # ── Índices de columnas ─────────────────────────────────────────
    def col_idx(name):
        try:
            return tw_hdrs.index(name)
        except ValueError:
            return None

    IDX_RUT   = col_idx('RUT')
    IDX_FICHA = col_idx('FICHA')
    IDX_DIAS_LIC  = col_idx('DIAS LICENCIA')
    IDX_DIAS_TRAB = col_idx('DIAS TRABAJADOS')
    IDX_FONASA    = col_idx('FONASA')
    IDX_ISAPRE    = col_idx('ISAPRE')
    IDX_AFP       = col_idx('AFP')
    IDX_LIQUIDO   = col_idx('LIQUIDO')
    IDX_SUELDO     = col_idx('SUELDO DEL MES')
    IDX_INST_SALUD  = col_idx('INST SALUD')
    IDX_NOMBRE_AFP  = col_idx('NOMBRE AFP')
    IDX_VAC      = col_idx('SUELDO POR VACACIONES')
    IDX_IMP1      = col_idx('IMPUESTO UNICO')
    IDX_IMP2      = col_idx('IMPUESTO UNICO DOBLE CONTRATO')
    IDX_SEG_SES   = col_idx('SEG SES TRAB')

    APV_COLS = [i for c, i in [
        ('APV',        col_idx('APV')),
        ('APV REG (A)', col_idx('APV REG (A)')),
        ('OTROS APV',  col_idx('OTROS APV')),
    ] if i is not None]

    haber_afecto_cols    = []
    haber_exento_cols    = []
    descuento_legal_cols = []
    descuento_cols       = []
    aporte_emp_cols      = []

    for col_name, eq in equiv.items():
        ci = col_idx(col_name)
        if ci is None or eq['id_concepto'] == 'No aplica concepto':
            continue
        tipo = eq['tipo']
        if tipo == 'Haber afecto':
            haber_afecto_cols.append((col_name, ci, eq))
        elif tipo == 'Haber exento':
            haber_exento_cols.append((col_name, ci, eq))
        elif tipo == 'Descuento Legal' and col_name not in DESC_LEGAL_MANUALES:
            descuento_legal_cols.append((col_name, ci, eq))
        elif tipo == 'Descuento' and col_name not in DESC_LEGAL_MANUALES:
            descuento_cols.append((col_name, ci, eq))
        elif tipo == 'Aporte Empleador' and col_name not in DESC_LEGAL_MANUALES:
            aporte_emp_cols.append((col_name, ci, eq))

    # ── Procesar empleados ──────────────────────────────────────────
    output_rows  = [OUTPUT_HEADERS]
    n_procesados = 0
    n_filas_gen  = [0]
    advertencias = []
    total        = len(tw_data)

    def safe_val(row, idx):
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    log("Procesando empleados…")

    for fila_n, row in enumerate(tw_data, start=9):

        ficha_raw = safe_val(row, IDX_FICHA)
        if ficha_raw is None or str(ficha_raw).strip() == '':
            continue

        ficha = str(ficha_raw).strip()
        rut   = formatear_rut(safe_val(row, IDX_RUT))

        dias_lic  = n(safe_val(row, IDX_DIAS_LIC))
        dias_trab = n(safe_val(row, IDX_DIAS_TRAB))

        emp = empleados.get(ficha, {})
        if not emp:
            advertencias.append(f"Fila {fila_n}: FICHA '{ficha}' no encontrada en empleadostw")

        id_empresa    = emp.get('Id empresa', '') or ''
        horas_sem     = n(emp.get('horasSema', 42))
        base_contrato = n(emp.get('Base contrato', 0))
        n_contrato    = emp.get('N° Contrato', '') if emp else ''
        jornada       = 'P' if horas_sem < 31 else 'C'

        # Sumas de haberes
        suma_afectos = 0.0
        for col_name, ci, eq in haber_afecto_cols:
            v = n(safe_val(row, ci))
            suma_afectos += (-v if col_name in HABERES_NEGATIVOS else v)

        suma_exentos = 0.0
        for col_name, ci, eq in haber_exento_cols:
            v = n(safe_val(row, ci))
            if col_name not in EXCLUIR_DE_EXENTOS:
                suma_exentos += v

        base_afp = max(min(suma_afectos, tope_afp) if tope_afp > 0 else suma_afectos, 0)
        base_ces = max(min(suma_afectos, tope_ces) if tope_ces > 0 else suma_afectos, 0)
        suma_afectos_pos = max(suma_afectos, 0)

        v_afp_col    = n(safe_val(row, IDX_AFP))
        v_seg_ses    = n(safe_val(row, IDX_SEG_SES))
        v_apv_sum    = sum(n(safe_val(row, i)) for i in APV_COLS)
        v_fonasa     = n(safe_val(row, IDX_FONASA))
        v_isapre_col = n(safe_val(row, IDX_ISAPRE))
        salud_total  = v_fonasa + v_isapre_col
        rebaja_salud = min(salud_total, tope_salud) if tope_salud > 0 else salud_total

        total_rebajas_llss = v_afp_col + v_apv_sum + v_seg_ses + rebaja_salud
        base_imp = max(suma_afectos - total_rebajas_llss, 0)
        monto_isapre = v_fonasa if v_fonasa != 0 else v_isapre_col

        def fila(id_concepto, monto, afecto_val=0, id_inst=None, cotiz=None,
                 total_reb=0, rentas_ng=0, monto_init=0):
            output_rows.append([
                periodo,
                rut,
                n_contrato,
                id_concepto,
                round(monto) if isinstance(monto, float) else monto,
                round(afecto_val),
                id_inst,
                cotiz,
                int(dias_lic),
                int(dias_trab),
                periodo,
                id_empresa,
                round(total_reb),
                round(rentas_ng),
                0,
                jornada,
                0,
                round(monto_init),
                1,
                0,
                0,
            ])
            n_filas_gen[0] += 1

        # sueldoBase = SUELDO DEL MES + SUELDO POR VACACIONES
        v_sueldo      = n(safe_val(row, IDX_SUELDO))
        v_vac         = n(safe_val(row, IDX_VAC))
        v_sueldo_base = v_sueldo + v_vac
        if v_sueldo_base != 0:
            fila('sueldoBase', v_sueldo_base, monto_init=base_contrato)

        # sueldoBaseVac — solo si SUELDO POR VACACIONES > 0
        if v_vac > 0:
            fila('sueldoBaseVac', v_vac)

        # sueldoBasesinVac — valor de SUELDO DEL MES
        if v_sueldo != 0:
            fila('sueldoBasesinVac', v_sueldo)

        # Haberes afectos agrupados
        hab_afecto_agrup = {}
        for col_name, ci, eq in haber_afecto_cols:
            if col_name in ('SUELDO DEL MES', 'SUELDO POR VACACIONES'):
                continue
            id_c = eq['id_concepto']
            v = n(safe_val(row, ci))
            monto_v = -v if col_name in HABERES_NEGATIVOS else v
            hab_afecto_agrup[id_c] = hab_afecto_agrup.get(id_c, 0) + monto_v
        for id_c, monto_v in hab_afecto_agrup.items():
            if monto_v != 0:
                fila(id_c, monto_v)

        # Haberes exentos agrupados
        hab_exento_agrup = {}
        for col_name, ci, eq in haber_exento_cols:
            id_c = eq['id_concepto']
            v = n(safe_val(row, ci))
            hab_exento_agrup[id_c] = hab_exento_agrup.get(id_c, 0) + v
        for id_c, monto_v in hab_exento_agrup.items():
            if monto_v != 0:
                fila(id_c, monto_v)

        # AFP
        if v_afp_col != 0:
            inst  = get_institucion('afp', emp, row, desde_ago)
            cotiz = get_cotizacion('afp', emp, base_afp, periodo, desde_ago,
                                   sis_tasa, aporte_afp, seg_vida, aporte_bac)
            fila('afp', v_afp_col, base_afp, inst, cotiz)

        # isapre (siempre) — cotización = mismo valor que monto
        inst_isapre = get_institucion('isapre', emp, row, desde_ago)
        fila('isapre', monto_isapre, base_afp, inst_isapre, monto_isapre)

        # cesEmpleado (siempre)
        inst_ces = get_institucion('cesEmpleado', emp, row, desde_ago)
        fila('cesEmpleado', v_seg_ses, base_ces, inst_ces, 0.6)

        # impuesto (siempre)
        v_imp1 = n(safe_val(row, IDX_IMP1))
        v_imp2 = n(safe_val(row, IDX_IMP2))
        fila('impuesto', v_imp1 + v_imp2, base_imp, 'Impuesto', None,
             total_reb=total_rebajas_llss, rentas_ng=suma_exentos)

        # APV
        if v_apv_sum != 0:
            fila('apvi', v_apv_sum)

        # Descuentos legales
        desc_legal_agrup = {}
        for col_name, ci, eq in descuento_legal_cols:
            id_c = eq['id_concepto']
            v = n(safe_val(row, ci))
            desc_legal_agrup[id_c] = desc_legal_agrup.get(id_c, 0) + v
        for id_c, monto_v in desc_legal_agrup.items():
            if monto_v != 0:
                afecto_v = get_afecto(id_c, base_afp, base_ces, base_imp, suma_afectos_pos)
                inst  = get_institucion(id_c, emp, row, desde_ago)
                cotiz = get_cotizacion(id_c, emp, afecto_v, periodo, desde_ago,
                                       sis_tasa, aporte_afp, seg_vida, aporte_bac)
                fila(id_c, monto_v, afecto_v, inst, cotiz)

        # Descuentos normales
        desc_agrup = {}
        for col_name, ci, eq in descuento_cols:
            id_c = eq['id_concepto']
            v = n(safe_val(row, ci))
            desc_agrup[id_c] = desc_agrup.get(id_c, 0) + v
        for id_c, monto_v in desc_agrup.items():
            if monto_v != 0:
                inst = get_institucion(id_c, emp, row, desde_ago)
                fila(id_c, monto_v, id_inst=inst)

        # Aportes empleador
        aporte_agrup = {}
        for col_name, ci, eq in aporte_emp_cols:
            id_c = eq['id_concepto']
            v = n(safe_val(row, ci))
            aporte_agrup[id_c] = aporte_agrup.get(id_c, 0) + v
        for id_c, monto_v in aporte_agrup.items():
            if monto_v != 0:
                afecto_v = get_afecto(id_c, base_afp, base_ces, base_imp, suma_afectos_pos)
                inst  = get_institucion(id_c, emp, row, desde_ago)
                cotiz = get_cotizacion(id_c, emp, afecto_v, periodo, desde_ago,
                                       sis_tasa, aporte_afp, seg_vida, aporte_bac)
                fila(id_c, monto_v, afecto_v, inst, cotiz)

        # totalesEmpl (LIQUIDO)
        v_liq = n(safe_val(row, IDX_LIQUIDO))
        cotiz_tot = get_cotizacion('totalesEmpl', emp, suma_afectos_pos, periodo,
                                   desde_ago, sis_tasa, aporte_afp, seg_vida, aporte_bac)
        fila('totalesEmpl', v_liq, suma_afectos_pos, None, cotiz_tot)

        # licenciaDias
        if dias_lic > 0:
            fila('licenciaDias', int(dias_lic))

        # cajaComp — si INST SALUD del archivo de entrada está en blanco o vacío
        inst_salud_tw = str(safe_val(row, IDX_INST_SALUD) or '').strip()
        if inst_salud_tw == '' and aporte_ccaf > 0:
            monto_ccaf = suma_afectos * (aporte_ccaf / 100)
            if monto_ccaf != 0:
                id_ccaf = emp.get('Id CCAF') or ''
                fila('cajaComp', monto_ccaf, id_inst=id_ccaf)

        n_procesados += 1
        if progress_callback and total > 0:
            progress_callback(n_procesados / total)

    stats = {
        'periodo': periodo,
        'empleados': n_procesados,
        'filas': n_filas_gen[0],
        'advertencias': len(advertencias),
    }
    return output_rows, advertencias, stats


def generar_excel(output_rows):
    """Genera el archivo Excel en memoria y retorna bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Liquidaciones'
    for row in output_rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────
#  INTERFAZ STREAMLIT
# ─────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="TeamWork → Rex+",
    page_icon="💼",
    layout="wide",
)

st.title("💼 TeamWork → Rex+")
st.caption("Transforma liquidaciones de TeamWork al formato de importación Rex+")

st.divider()

# ── Columnas para los upload widgets ───────────────────────────────
col1, col2 = st.columns(2)

with col1:
    st.subheader("Archivos de datos")
    tw_file     = st.file_uploader("📄 tw.xlsx  *(nómina principal)*",
                                   type=["xlsx"], key="tw")
    emp_file    = st.file_uploader("👥 empleadostw.xlsx  *(maestro de empleados)*",
                                   type=["xlsx"], key="emp")
    equiv_file  = st.file_uploader("🔄 Equivalencias Tw.xlsx  *(mapeo de conceptos)*",
                                   type=["xlsx"], key="equiv")

with col2:
    st.subheader("Archivos de parámetros")
    params_file = st.file_uploader("⚙️ parametrosMesuales.xlsx  *(topes y tasas)*",
                                   type=["xlsx"], key="params")
    cot_file    = st.file_uploader("📊 cot_afp_hist.xlsx  *(cotizaciones AFP)*",
                                   type=["xlsx"], key="cot")
    asig_file   = st.file_uploader("🏦 Asig Inst LD.xlsx  *(instituciones por concepto)*",
                                   type=["xlsx"], key="asig")

st.divider()

# ── Botón Procesar ──────────────────────────────────────────────
archivos_ok = all([tw_file, emp_file, equiv_file, params_file, cot_file])

if not archivos_ok:
    faltantes = []
    if not tw_file:     faltantes.append("tw.xlsx")
    if not emp_file:    faltantes.append("empleadostw.xlsx")
    if not equiv_file:  faltantes.append("Equivalencias Tw.xlsx")
    if not params_file: faltantes.append("parametrosMesuales.xlsx")
    if not cot_file:    faltantes.append("cot_afp_hist.xlsx")
    if not asig_file:   faltantes.append("Asig Inst LD.xlsx")
    st.info(f"Carga los 6 archivos para habilitar el proceso. Faltan: **{', '.join(faltantes)}**")

btn_procesar = st.button("🚀 Procesar", type="primary", disabled=not archivos_ok)

if btn_procesar and archivos_ok:
    log_lines  = []
    log_box    = st.empty()
    prog_bar   = st.progress(0.0, text="Iniciando…")
    status_msg = st.empty()

    def log_callback(msg):
        log_lines.append(msg)
        log_box.code('\n'.join(log_lines[-30:]))   # últimas 30 líneas

    def progress_callback(frac):
        prog_bar.progress(min(frac, 1.0), text=f"Procesando empleados… {frac*100:.1f}%")

    try:
        output_rows, advertencias, stats = procesar(
            tw_bytes    = tw_file.read(),
            equiv_bytes = equiv_file.read(),
            emp_bytes   = emp_file.read(),
            params_bytes= params_file.read(),
            cot_bytes   = cot_file.read(),
            asig_bytes  = asig_file.read() if asig_file else None,
            progress_callback = progress_callback,
            log_callback      = log_callback,
        )

        prog_bar.progress(1.0, text="Generando archivo de salida…")
        excel_bytes = generar_excel(output_rows)
        prog_bar.empty()
        log_box.empty()

        # ── Resultados ──────────────────────────────────────────────
        st.success("✅ Proceso completado")

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Período", stats['periodo'])
        m2.metric("Empleados procesados", f"{stats['empleados']:,}")
        m3.metric("Filas generadas", f"{stats['filas']:,}")
        m4.metric("Advertencias", stats['advertencias'])

        nombre_salida = f"salida_rex_{stats['periodo']}.xlsx"
        st.download_button(
            label     = "⬇️ Descargar salida_rex.xlsx",
            data      = excel_bytes,
            file_name = nombre_salida,
            mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type      = "primary",
        )

        if advertencias:
            with st.expander(f"⚠️ Ver advertencias ({len(advertencias)})"):
                for a in advertencias[:200]:
                    st.text(a)
                if len(advertencias) > 200:
                    st.caption(f"… y {len(advertencias) - 200} advertencias más.")

    except Exception as e:
        prog_bar.empty()
        st.error(f"❌ Error durante el proceso: {e}")
        with st.expander("Detalle del error"):
            import traceback
            st.code(traceback.format_exc())
