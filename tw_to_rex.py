#!/usr/bin/env python3
"""
Script de transformación TeamWork → Rex+
Lee tw.xlsx y archivos de referencia, genera salida_rex.xlsx con la
estructura de importación de liquidaciones de Rex+.

Reglas incorporadas (validadas en sesiones previas):
  - DESCTO. HORAS ATRASO, DESCTO. HORAS ATRASO PT,
    HORAS NO TRABAJADAS $, DESC. PAGO EN EXC. IMPONIBLE → suman negativo
  - SOBREGIRO → se incluye como fila pero NO entra en rentas no gravadas
  - Tope AFP aplica a base_afp cuando corresponde
  - SIS: antes de Aug-2026 → Id AFP; desde Aug-2026 → seguridadsocial
"""

import openpyxl
import re
import sys
import os

# ─────────────────────────────────────────────────────────────────
#  PARÁMETROS DE ARCHIVOS (rutas relativas al directorio de trabajo)
# ─────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

TW_FILE       = os.path.join(BASE_DIR, "tw.xlsx")
EQUIV_FILE    = os.path.join(BASE_DIR, "Equivalencias Tw.xlsx")
EMPLEADOS_FILE= os.path.join(BASE_DIR, "empleadostw.xlsx")
PARAMS_FILE   = os.path.join(BASE_DIR, "parametrosMesuales.xlsx")
COT_AFP_FILE  = os.path.join(BASE_DIR, "cot_afp_hist.xlsx")
ASIG_FILE     = os.path.join(BASE_DIR, "Asig Inst LD.xlsx")
OUTPUT_FILE   = os.path.join(BASE_DIR, "salida_rex.xlsx")

# Columnas con monto que en realidad restan (haberes negativos)
HABERES_NEGATIVOS = {
    'DESCTO. HORAS ATRASO',
    'DESCTO. HORAS ATRASO PT',
    'HORAS NO TRABAJADAS $',
    'DESC. PAGO EN EXC. IMPONIBLE',
}

# Columna que NO entra en rentas no gravadas (exentos) para impuesto
EXCLUIR_DE_EXENTOS = set()   # SOBREGIRO (compensaSobre) sí entra en rentas no gravadas

# Conceptos que siempre generan fila aunque monto sea 0
SIEMPRE_GENERAR = {'impuesto', 'cesEmpleado'}

# Meses en español → número
MESES_ES = {
    'ENERO':'01','FEBRERO':'02','MARZO':'03','ABRIL':'04',
    'MAYO':'05','JUNIO':'06','JULIO':'07','AGOSTO':'08',
    'SEPTIEMBRE':'09','OCTUBRE':'10','NOVIEMBRE':'11','DICIEMBRE':'12'
}

# Conceptos que usan base AFP (min(haberes_afectos, tope_afp))
CONCEPTOS_BASE_AFP = {
    'afp','afpAhor','reliquidaAfp','aporteAFPemp','reliquidaAporteAFP',
    'aporteFAPPBAC','aporteFAPPCEV','mutual','reliquidaMutual',
    'sis','trabajoPesa','reliquidaTrabPesa','trabajoPesaEmpl','reliquidaTrabEmpl',
    'isapre','reliquidaIsapre',
}
# Conceptos que usan base CES (min(haberes_afectos, tope_ces))
CONCEPTOS_BASE_CES = {
    'cesEmpleado','cesAporteSol','cesAporteCi',
    'reliquidaCesSol','reliquidaCesCi','reliquidaCesEmpl','solidarioremu',
}
# Conceptos que usan base impuesto (haberes_afectos - rebajas_llss)
CONCEPTOS_BASE_IMP = {
    'impuesto','impuestoAgricola','reliquidaImpuesto','imptoindMi',
}
# totalesEmpl → suma_haberes_afectos sin tope
CONCEPTOS_BASE_TOT = {'totalesEmpl'}


# ─────────────────────────────────────────────────────────────────
#  FUNCIONES AUXILIARES
# ─────────────────────────────────────────────────────────────────

def n(v):
    """Convierte a float, 0 si None o error."""
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(',', '.').strip())
    except Exception:
        return 0.0


def formatear_rut(raw):
    """'020.238.296-7' → '20238296-7'"""
    if not raw:
        return ''
    s = str(raw).strip().replace('.', '')
    partes = s.split('-')
    if len(partes) == 2:
        return f"{partes[0].lstrip('0')}-{partes[1]}"
    return s


def parsear_periodo(texto):
    """'Mes a procesar: JUNIO 2026' → '2026-06'"""
    upper = texto.upper() if texto else ''
    for nombre, num in MESES_ES.items():
        if nombre in upper:
            m = re.search(r'(\d{4})', texto)
            if m:
                return f"{m.group(1)}-{num}"
    return None


def desde_agosto_2026(periodo):
    """True si el período es >= 2026-08"""
    try:
        y, mo = map(int, periodo.split('-'))
        return (y > 2026) or (y == 2026 and mo >= 8)
    except Exception:
        return False


def get_afecto(id_concepto, base_afp, base_ces, base_imp, suma_afectos):
    """Retorna el valor de Afecto según tipo de concepto."""
    if id_concepto in CONCEPTOS_BASE_AFP:
        return round(base_afp)
    if id_concepto in CONCEPTOS_BASE_CES:
        return round(base_ces)
    if id_concepto in CONCEPTOS_BASE_IMP:
        return round(base_imp)
    if id_concepto in CONCEPTOS_BASE_TOT:
        return round(suma_afectos)
    return 0


# ─────────────────────────────────────────────────────────────────
#  CARGAR ARCHIVOS DE REFERENCIA
# ─────────────────────────────────────────────────────────────────

print("=" * 60)
print("Cargando archivos de referencia…")
print("=" * 60)

# ── Equivalencias ──
wb = openpyxl.load_workbook(EQUIV_FILE, data_only=True, read_only=True)
ws = wb.active
equiv = {}   # col_name → {id_concepto, tipo, obs}
for row in ws.iter_rows(values_only=True):
    if not row[0] or row[0] == 'Nombre columna':
        continue
    col_name   = str(row[0]).strip()
    id_concepto= str(row[1]).strip() if row[1] else ''
    tipo       = str(row[2]).strip() if row[2] else ''
    obs        = str(row[3]).strip() if row[3] else ''
    equiv[col_name] = {'id_concepto': id_concepto, 'tipo': tipo, 'obs': obs}
wb.close()
print(f"  Equivalencias: {len(equiv)} filas")

# ── Empleados TW ──
wb = openpyxl.load_workbook(EMPLEADOS_FILE, data_only=True, read_only=True)
ws = wb.active
rows_emp = list(ws.iter_rows(values_only=True))
wb.close()
emp_headers = [str(h).strip() if h else '' for h in rows_emp[1]]
empleados = {}  # Nombre del contrato (str) → dict   [join por FICHA]
for row in rows_emp[2:]:
    if not row[0]:
        continue
    d = {emp_headers[i]: row[i] for i in range(len(emp_headers))}
    key = str(d.get('Nombre del contrato', '')).strip()
    if key:
        empleados[key] = d
print(f"  Empleados cargados: {len(empleados)}")

# ── Parámetros Mensuales ──
wb = openpyxl.load_workbook(PARAMS_FILE, data_only=True, read_only=True)
ws = wb.active
rows_pm = list(ws.iter_rows(values_only=True))
wb.close()
pm_hdrs = [str(h).strip() if h else '' for h in rows_pm[0]]
params = {}  # periodo → dict
for row in rows_pm[1:]:
    if not row[0]:
        continue
    periodo_key = str(row[0]).strip()
    params[periodo_key] = {pm_hdrs[i]: row[i] for i in range(len(pm_hdrs))}
print(f"  Períodos en parámetros: {len(params)}")

# ── Cotizaciones AFP ──
wb = openpyxl.load_workbook(COT_AFP_FILE, data_only=True, read_only=True)
ws = wb.active
cot_afp = {}  # 'periodo+afp_id' → cot_hist_afp (decimal)
for row in ws.iter_rows(values_only=True):
    if not row[0] or row[0] == 'id_afp_hist':
        continue
    cot_afp[str(row[0]).strip()] = n(row[4])
wb.close()
print(f"  Cotizaciones AFP: {len(cot_afp)} registros")

# ── Asignación de Instituciones ──
wb = openpyxl.load_workbook(ASIG_FILE, data_only=True, read_only=True)
ws = wb.active
rows_ai = list(ws.iter_rows(values_only=True))
wb.close()
asig = {}  # concepto → {hasta_jul, desde_ago}
for row in rows_ai[1:]:   # fila 1 es header de período
    if not row[0]:
        continue
    asig[str(row[0]).strip()] = {
        'hasta_jul': str(row[1]).strip() if row[1] else '',
        'desde_ago': str(row[2]).strip() if row[2] else '',
    }
print(f"  Reglas de institución: {len(asig)} conceptos")


# ─────────────────────────────────────────────────────────────────
#  FUNCIONES QUE USAN LAS TABLAS CARGADAS
# ─────────────────────────────────────────────────────────────────

def get_institucion(concepto, emp, desde_ago):
    """Devuelve el Id de institución según concepto y período."""
    regla = asig.get(concepto, {})
    col = regla.get('desde_ago' if desde_ago else 'hasta_jul', '')
    if col == 'Id afp':
        return emp.get('Id Afp') or ''
    if col == 'Id Salud':
        return emp.get('Id Salud') or ''
    if col == 'Id Mutual':
        return emp.get('Id Mutual') or ''
    if col == 'Id CCAF':
        return emp.get('Id CCAF') or ''
    if col == 'seguridadsocial':
        return 'seguridadsocial'
    if col == 'Impuesto':
        return 'Impuesto'
    if col == 'No aplica':
        return None
    return None


def get_cotizacion(concepto, emp, afecto, periodo, desde_ago, sis_tasa,
                   aporte_afp, seg_vida, aporte_bac):
    """Devuelve la cotización de jubilación para el concepto."""
    id_afp = emp.get('Id Afp', '') or ''
    if concepto == 'afp':
        key = f"{periodo}{id_afp}"
        cot = cot_afp.get(key, 0)
        # Fallback: buscar el período anterior si no existe el actual
        if not cot and '-' in periodo:
            y, mo = int(periodo[:4]), int(periodo[5:])
            mo -= 1
            if mo == 0:
                mo, y = 12, y - 1
            key_prev = f"{y:04d}-{mo:02d}{id_afp}"
            cot = cot_afp.get(key_prev, 0)
        return round(cot * 100, 4) if cot else None
    if concepto == 'isapre':
        return afecto   # = base_afp
    if concepto == 'mutual':
        return n(emp.get('% Mutual', 0))
    if concepto == 'sis':
        return sis_tasa
    if concepto == 'aporteAFPemp':
        return aporte_afp
    if concepto == 'aporteFAPPCEV':
        return seg_vida
    if concepto == 'aporteFAPPBAC':
        return aporte_bac
    if concepto == 'totalesEmpl':
        return afecto   # = suma_afectos
    if concepto == 'cesEmpleado':
        return 0.6
    return None


# ─────────────────────────────────────────────────────────────────
#  LEER ARCHIVO TW
# ─────────────────────────────────────────────────────────────────

print("\nCargando tw.xlsx…")
wb = openpyxl.load_workbook(TW_FILE, data_only=True, read_only=True)
ws = wb.active
tw_rows_all = list(ws.iter_rows(values_only=True))
wb.close()

# Período: fila 2 (índice 1)
periodo = parsear_periodo(str(tw_rows_all[1][0]))
if not periodo:
    print("ERROR: No se pudo detectar el período en la fila 2 de tw.xlsx")
    sys.exit(1)
print(f"  Período: {periodo}")

# Headers: fila 8 (índice 7)
tw_hdrs = [str(h).strip() if h is not None else '' for h in tw_rows_all[7]]

# Datos: fila 9 en adelante (índice 8+)
tw_data = tw_rows_all[8:]
print(f"  Filas de datos: {len(tw_data)}")

# ── Parámetros del período ──
pm = params.get(periodo, {})
if not pm:
    print(f"  ADVERTENCIA: Sin parámetros para {periodo}")

tope_afp   = n(pm.get('topeImp_pesos_afp', 0))
tope_ces   = n(pm.get('topeCes_pesos', 0))
tope_salud = n(pm.get('topeSalud_pesos', 0))
sis_tasa   = n(pm.get('sis', 0))
aporte_afp = n(pm.get('Aporte AFP', 0))
seg_vida   = n(pm.get('Seg Social Exp vida', 0))
aporte_bac = n(pm.get('aporteFAPPBAC', 0))

desde_ago = desde_agosto_2026(periodo)
print(f"  Tope AFP : {tope_afp:>12,.0f}")
print(f"  Tope CES : {tope_ces:>12,.0f}")
print(f"  Tope Salud: {tope_salud:>11,.0f}")
print(f"  SIS tasa  : {sis_tasa}")
print(f"  Desde Ago 2026: {desde_ago}")


# ── Índices de columnas importantes ──
def col_idx(name):
    try:
        return tw_hdrs.index(name)
    except ValueError:
        return None

IDX_RUT       = col_idx('RUT')
IDX_FICHA     = col_idx('FICHA')
IDX_DIAS_LIC  = col_idx('DIAS LICENCIA')
IDX_DIAS_TRAB = col_idx('DIAS TRABAJADOS')
IDX_FONASA    = col_idx('FONASA')
IDX_ISAPRE    = col_idx('ISAPRE')
IDX_AFP       = col_idx('AFP')
IDX_LIQUIDO   = col_idx('LIQUIDO')
IDX_SUELDO    = col_idx('SUELDO DEL MES')
IDX_IMP1      = col_idx('IMPUESTO UNICO')
IDX_IMP2      = col_idx('IMPUESTO UNICO DOBLE CONTRATO')
IDX_SEG_SES   = col_idx('SEG SES TRAB')

APV_COLS = [i for c, i in [('APV', col_idx('APV')),
                            ('APV REG (A)', col_idx('APV REG (A)')),
                            ('OTROS APV', col_idx('OTROS APV'))]
            if i is not None]

# Columnas de haberes afectos (para calcular suma_afectos)
haber_afecto_cols = []   # (col_name, idx, eq)
haber_exento_cols = []   # (col_name, idx, eq)

for col_name, eq in equiv.items():
    ci = col_idx(col_name)
    if ci is None:
        continue
    if eq['id_concepto'] == 'No aplica concepto':
        continue
    if eq['tipo'] == 'Haber afecto':
        haber_afecto_cols.append((col_name, ci, eq))
    elif eq['tipo'] == 'Haber exento':
        haber_exento_cols.append((col_name, ci, eq))

# Columnas de descuentos legales (excluye los manejados a mano)
DESC_LEGAL_MANUALES = {'AFP', 'FONASA', 'ISAPRE', 'IMPUESTO UNICO',
                       'IMPUESTO UNICO DOBLE CONTRATO', 'SEG SES TRAB',
                       'APV', 'APV REG (A)', 'OTROS APV', 'LIQUIDO',
                       'SUELDO DEL MES'}

descuento_legal_cols = []  # (col_name, idx, eq)
descuento_cols       = []  # (col_name, idx, eq)
aporte_emp_cols      = []  # (col_name, idx, eq)

for col_name, eq in equiv.items():
    ci = col_idx(col_name)
    if ci is None or col_name in DESC_LEGAL_MANUALES:
        continue
    if eq['id_concepto'] == 'No aplica concepto':
        continue
    tipo = eq['tipo']
    if tipo == 'Descuento Legal':
        descuento_legal_cols.append((col_name, ci, eq))
    elif tipo == 'Descuento':
        descuento_cols.append((col_name, ci, eq))
    elif tipo == 'Aporte Empleador':
        aporte_emp_cols.append((col_name, ci, eq))

print(f"\n  Haberes afectos mapeados  : {len(haber_afecto_cols)}")
print(f"  Haberes exentos mapeados  : {len(haber_exento_cols)}")
print(f"  Descuentos legales mapeados: {len(descuento_legal_cols)}")
print(f"  Descuentos mapeados       : {len(descuento_cols)}")
print(f"  Aportes empleador mapeados: {len(aporte_emp_cols)}")


# ─────────────────────────────────────────────────────────────────
#  GENERAR FILAS DE SALIDA
# ─────────────────────────────────────────────────────────────────

OUTPUT_HEADERS = [
    'Fecha de proceso', 'Id empleado', 'Número de contrato', 'Id del concepto',
    'Monto del concepto', 'Afecto', 'Id de institución', 'Cotización de jubilación',
    'Días de licencias', 'Días trabajados', 'Fecha de aplicación', 'Empresa',
    'Total de rebajas por LLSS', 'Rentas no gravadas', 'Rebaja por zona extrema',
    'Jornada', 'Días de vacaciones', 'Monto Init', 'Fase', 'Parcial7', 'Parcial8'
]

output_rows  = [OUTPUT_HEADERS]
n_procesados = 0
n_filas_gen  = [0]   # lista mutable para que la función anidada pueda modificarlo
advertencias = []

def safe_val(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


print(f"\nProcesando {len(tw_data)} empleados…")

for fila_n, row in enumerate(tw_data, start=9):

    ficha_raw = safe_val(row, IDX_FICHA)
    if ficha_raw is None or str(ficha_raw).strip() == '':
        continue

    ficha = str(ficha_raw).strip()
    rut   = formatear_rut(safe_val(row, IDX_RUT))

    dias_lic  = n(safe_val(row, IDX_DIAS_LIC))
    dias_trab = n(safe_val(row, IDX_DIAS_TRAB))

    # Datos del empleado
    emp = empleados.get(ficha, {})
    if not emp:
        advertencias.append(f"Fila {fila_n}: FICHA '{ficha}' no encontrada en empleadostw")

    id_empresa   = emp.get('Id empresa', '') or ''
    horas_sem    = n(emp.get('horasSema', 42))
    base_contrato= n(emp.get('Base contrato', 0))
    n_contrato   = emp.get('N° Contrato', '') if emp else ''
    jornada      = 'P' if horas_sem < 31 else 'C'

    # ── Calcular sumas de haberes ──────────────────────────────────
    suma_afectos = 0.0
    for col_name, ci, eq in haber_afecto_cols:
        v = n(safe_val(row, ci))
        if col_name in HABERES_NEGATIVOS:
            suma_afectos -= v
        else:
            suma_afectos += v

    suma_exentos = 0.0   # para rentas no gravadas (excluye SOBREGIRO)
    for col_name, ci, eq in haber_exento_cols:
        v = n(safe_val(row, ci))
        if col_name not in EXCLUIR_DE_EXENTOS:
            suma_exentos += v

    # Topes
    base_afp = min(suma_afectos, tope_afp) if tope_afp > 0 else suma_afectos
    base_ces = min(suma_afectos, tope_ces) if tope_ces > 0 else suma_afectos
    base_afp = max(base_afp, 0)
    base_ces = max(base_ces, 0)
    suma_afectos_pos = max(suma_afectos, 0)

    # ── Valores puntuales para LLSS ───────────────────────────────
    v_afp_col  = n(safe_val(row, IDX_AFP))
    v_seg_ses  = n(safe_val(row, IDX_SEG_SES))
    v_apv_sum  = sum(n(safe_val(row, i)) for i in APV_COLS)
    v_fonasa   = n(safe_val(row, IDX_FONASA))
    v_isapre_col = n(safe_val(row, IDX_ISAPRE))
    salud_total  = v_fonasa + v_isapre_col
    rebaja_salud = min(salud_total, tope_salud) if tope_salud > 0 else salud_total

    total_rebajas_llss = v_afp_col + v_apv_sum + v_seg_ses + rebaja_salud
    base_imp = suma_afectos - total_rebajas_llss
    base_imp = max(base_imp, 0)

    # Monto isapre (el que sea distinto de 0; preferir FONASA)
    monto_isapre = v_fonasa if v_fonasa != 0 else v_isapre_col

    # ── Función auxiliar para agregar fila ─────────────────────────
    def fila(id_concepto, monto, afecto_val=0, id_inst=None, cotiz=None,
             total_reb=0, rentas_ng=0, monto_init=0):
        output_rows.append([
            periodo,            # Fecha de proceso
            rut,                # Id empleado
            n_contrato,         # Número de contrato
            id_concepto,        # Id del concepto
            round(monto) if isinstance(monto, float) else monto,
            round(afecto_val),  # Afecto
            id_inst,            # Id de institución
            cotiz,              # Cotización de jubilación
            int(dias_lic),      # Días de licencias
            int(dias_trab),     # Días trabajados
            periodo,            # Fecha de aplicación
            id_empresa,         # Empresa
            round(total_reb),   # Total de rebajas por LLSS
            round(rentas_ng),   # Rentas no gravadas
            0,                  # Rebaja por zona extrema
            jornada,            # Jornada
            0,                  # Días de vacaciones
            round(monto_init),  # Monto Init
            1,                  # Fase
            0,                  # Parcial7
            0,                  # Parcial8
        ])
        n_filas_gen[0] += 1

    # ── sueldoBase ─────────────────────────────────────────────────
    v_sueldo = n(safe_val(row, IDX_SUELDO))
    if v_sueldo != 0:
        fila('sueldoBase', v_sueldo, monto_init=base_contrato)

    # ── Haberes afectos (excepto sueldoBase que ya se agregó) ──────
    # Agrupar por id_concepto (pueden haber columnas múltiples → mismo concepto)
    hab_afecto_agrup = {}
    for col_name, ci, eq in haber_afecto_cols:
        if col_name == 'SUELDO DEL MES':
            continue
        id_c = eq['id_concepto']
        v = n(safe_val(row, ci))
        monto_v = -v if col_name in HABERES_NEGATIVOS else v
        hab_afecto_agrup[id_c] = hab_afecto_agrup.get(id_c, 0) + monto_v

    for id_c, monto_v in hab_afecto_agrup.items():
        if monto_v != 0:
            fila(id_c, monto_v)

    # ── Haberes exentos ────────────────────────────────────────────
    hab_exento_agrup = {}
    for col_name, ci, eq in haber_exento_cols:
        id_c = eq['id_concepto']
        v = n(safe_val(row, ci))
        hab_exento_agrup[id_c] = hab_exento_agrup.get(id_c, 0) + v

    for id_c, monto_v in hab_exento_agrup.items():
        if monto_v != 0:
            fila(id_c, monto_v)

    # ── AFP ────────────────────────────────────────────────────────
    if v_afp_col != 0:
        inst = get_institucion('afp', emp, desde_ago)
        cotiz = get_cotizacion('afp', emp, base_afp, periodo, desde_ago,
                               sis_tasa, aporte_afp, seg_vida, aporte_bac)
        fila('afp', v_afp_col, base_afp, inst, cotiz)

    # ── isapre (siempre generar) ───────────────────────────────────
    inst_isapre = get_institucion('isapre', emp, desde_ago)
    fila('isapre', monto_isapre, base_afp, inst_isapre, base_afp)

    # ── cesEmpleado (siempre generar) ─────────────────────────────
    inst_ces = get_institucion('cesEmpleado', emp, desde_ago)
    fila('cesEmpleado', v_seg_ses, base_ces, inst_ces, 0.6)

    # ── impuesto (siempre generar) ─────────────────────────────────
    v_imp1 = n(safe_val(row, IDX_IMP1))
    v_imp2 = n(safe_val(row, IDX_IMP2))
    monto_imp = v_imp1 + v_imp2
    fila('impuesto', monto_imp, base_imp, 'Impuesto', None,
         total_reb=total_rebajas_llss, rentas_ng=suma_exentos)

    # ── APV (apvi) ─────────────────────────────────────────────────
    if v_apv_sum != 0:
        fila('apvi', v_apv_sum)

    # ── Otros descuentos legales (PRESTAMO SOLIDARIO, AHORRO AFP, etc.) ──
    desc_legal_agrup = {}
    for col_name, ci, eq in descuento_legal_cols:
        id_c = eq['id_concepto']
        v = n(safe_val(row, ci))
        desc_legal_agrup[id_c] = desc_legal_agrup.get(id_c, 0) + v

    for id_c, monto_v in desc_legal_agrup.items():
        if monto_v != 0:
            afecto_v = get_afecto(id_c, base_afp, base_ces, base_imp, suma_afectos_pos)
            inst = get_institucion(id_c, emp, desde_ago)
            cotiz = get_cotizacion(id_c, emp, afecto_v, periodo, desde_ago,
                                   sis_tasa, aporte_afp, seg_vida, aporte_bac)
            fila(id_c, monto_v, afecto_v, inst, cotiz)

    # ── Descuentos normales ────────────────────────────────────────
    desc_agrup = {}
    for col_name, ci, eq in descuento_cols:
        id_c = eq['id_concepto']
        v = n(safe_val(row, ci))
        desc_agrup[id_c] = desc_agrup.get(id_c, 0) + v

    for id_c, monto_v in desc_agrup.items():
        if monto_v != 0:
            fila(id_c, monto_v)

    # ── Aportes empleador ─────────────────────────────────────────
    aporte_agrup = {}
    for col_name, ci, eq in aporte_emp_cols:
        id_c = eq['id_concepto']
        v = n(safe_val(row, ci))
        aporte_agrup[id_c] = aporte_agrup.get(id_c, 0) + v

    for id_c, monto_v in aporte_agrup.items():
        if monto_v != 0:
            afecto_v = get_afecto(id_c, base_afp, base_ces, base_imp, suma_afectos_pos)
            inst = get_institucion(id_c, emp, desde_ago)
            cotiz = get_cotizacion(id_c, emp, afecto_v, periodo, desde_ago,
                                   sis_tasa, aporte_afp, seg_vida, aporte_bac)
            fila(id_c, monto_v, afecto_v, inst, cotiz)

    # ── totalesEmpl (LIQUIDO) ──────────────────────────────────────
    v_liq = n(safe_val(row, IDX_LIQUIDO))
    cotiz_tot = get_cotizacion('totalesEmpl', emp, suma_afectos_pos, periodo,
                               desde_ago, sis_tasa, aporte_afp, seg_vida, aporte_bac)
    fila('totalesEmpl', v_liq, suma_afectos_pos, None, cotiz_tot)

    # ── licenciaDias (si hay días de licencia) ─────────────────────
    if dias_lic > 0:
        fila('licenciaDias', int(dias_lic))

    n_procesados += 1
    if n_procesados % 500 == 0:
        print(f"  … {n_procesados} empleados procesados ({n_filas_gen[0]} filas)")


print(f"\n  Empleados procesados : {n_procesados}")
print(f"  Filas de salida      : {n_filas_gen[0]}")
if advertencias:
    print(f"\n  ADVERTENCIAS ({min(len(advertencias), 20)} de {len(advertencias)}):")
    for a in advertencias[:20]:
        print(f"    {a}")
    if len(advertencias) > 20:
        print(f"    … y {len(advertencias)-20} más")


# ─────────────────────────────────────────────────────────────────
#  GUARDAR SALIDA
# ─────────────────────────────────────────────────────────────────

print(f"\nGuardando {OUTPUT_FILE} …")
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = 'Liquidaciones'
for r in output_rows:
    ws_out.append(r)
wb_out.save(OUTPUT_FILE)
print(f"  ✓ Guardado: {OUTPUT_FILE}")
print("  Proceso completado.")
